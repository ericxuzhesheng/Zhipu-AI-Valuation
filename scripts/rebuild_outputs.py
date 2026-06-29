from __future__ import annotations

from dataclasses import dataclass
from pathlib import Path
from typing import NamedTuple
import os

import matplotlib.pyplot as plt
import numpy as np
import openpyxl
import pandas as pd
from math import comb
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter

from event_panel import write_event_panel_outputs


ROOT = Path(__file__).resolve().parents[1]
FIGURES = ROOT / "figures"
EVENTSTUDY = ROOT / "eventstudy"
MODEL = ROOT / "model" / "valuation_model.xlsx"


WACC = 0.135
TERMINAL_G = 0.04
TAX_RATE = 0.15
SHARES_M = 435
USD_HKD = 7.8
NET_CASH_USDM = 550
REV_2025_USDM = 102
REV_2026_USDM = 200

# ---- Coordinated high-contrast palette: deep blue / deep red + orange & green accents ----
C_BLUE = "#1F4E79"    # deep blue   — peer series (MiniMax), main bars, history
C_RED = "#B22222"     # deep red    — protagonist (Zhipu), market, emphasis
C_ORANGE = "#E08214"  # deep orange — capability releases
C_GREEN = "#2E7D32"   # deep green  — index / flow events, secondary category
C_TEAL = "#2C7FB8"    # blue-teal   — tertiary / earlier history
C_GRAY = "#9AA0A6"    # neutral gray — non-event bars / muted reference
C_INK = "#222222"     # near-black  — base/zero lines, averages


@dataclass(frozen=True)
class Scenario:
    name: str
    probability: float
    growth_start_2027: float
    growth_end_2035: float
    terminal_margin: float
    sales_to_capital: float


SCENARIOS = [
    Scenario("Bear", 0.35, 0.40, 0.06, 0.18, 2.0),
    Scenario("Base", 0.45, 0.56, 0.08, 0.28, 1.8),
    Scenario("Bull", 0.20, 0.85, 0.12, 0.35, 1.6),
]


def write_csv(df: pd.DataFrame, path: Path) -> None:
    tmp_path = path.with_name(f"{path.stem}.tmp{path.suffix}")
    df.to_csv(tmp_path, index=False)
    os.replace(tmp_path, path)


def savefig(fig: plt.Figure, path: Path, **kwargs) -> None:
    tmp_path = path.with_name(f"{path.stem}.tmp{path.suffix}")
    fig.savefig(tmp_path, **kwargs)
    os.replace(tmp_path, path)


class MarketSnapshot(NamedTuple):
    date: pd.Timestamp
    price_hkd: float
    equity_value_usdm: float
    revenue_multiple: float
    ret_vs_ipo_pct: float
    ann_vol_pct: float


def _price_frame(filename: str, ipo_price: float) -> pd.DataFrame:
    df = pd.read_csv(ROOT / "data" / filename)
    df["date"] = pd.to_datetime(df["trade_date"].astype(str))
    df = df.sort_values("date").reset_index(drop=True)
    df["rebased"] = df["close"] / ipo_price * 100.0
    df["daily_return"] = df["close"].pct_change()
    return df


def market_snapshot() -> MarketSnapshot:
    zhipu = _price_frame("Zhipu_KnowledgeAtlas_daily.csv", 116.2)
    latest = zhipu.iloc[-1]
    price = float(latest["close"])
    equity_value_usdm = price * SHARES_M / USD_HKD
    revenue_multiple = equity_value_usdm / REV_2026_USDM
    ann_vol_pct = float(zhipu["daily_return"].dropna().std(ddof=1) * np.sqrt(252) * 100)
    return MarketSnapshot(
        date=pd.Timestamp(latest["date"]),
        price_hkd=price,
        equity_value_usdm=equity_value_usdm,
        revenue_multiple=revenue_multiple,
        ret_vs_ipo_pct=(price / 116.2 - 1) * 100,
        ann_vol_pct=ann_vol_pct,
    )


MARKET = market_snapshot()


def write_price_summary_csv() -> None:
    specs = [
        ("02513.HK", "Zhipu_KnowledgeAtlas", "Zhipu_KnowledgeAtlas_daily.csv", 116.2),
        ("00100.HK", "MiniMax", "MiniMax_daily.csv", 165.0),
        ("01956.HK", "WengeAI", "WengeAI_daily.csv", 60.7),
    ]
    rows = []
    for code, name, filename, ipo_price in specs:
        df = _price_frame(filename, ipo_price)
        latest = df.iloc[-1]
        daily_returns = df["daily_return"].dropna()
        ann_vol_pct = (
            float(daily_returns.std(ddof=1) * np.sqrt(252) * 100)
            if len(daily_returns) >= 2
            else 0.0
        )
        rows.append(
            {
                "code": code,
                "name": name,
                "ipo_price": ipo_price,
                "first_day_close": round(float(df.iloc[0]["close"]), 1),
                "latest_date": int(latest["date"].strftime("%Y%m%d")),
                "latest_close": round(float(latest["close"]), 1),
                "ret_vs_ipo": round((float(latest["close"]) / ipo_price - 1) * 100, 1),
                "high": round(float(df["close"].max()), 1),
                "low": round(float(df["close"].min()), 1),
                "ann_vol_pct": round(ann_vol_pct, 1),
                "n_days": int(len(df)),
            }
        )
    write_csv(pd.DataFrame(rows), ROOT / "data" / "price_summary.csv")


def project_scenario(
    scenario: Scenario,
    *,
    wacc: float = WACC,
    terminal_margin: float | None = None,
    growth_start_2027: float | None = None,
) -> dict:
    terminal_margin = scenario.terminal_margin if terminal_margin is None else terminal_margin
    growth_start_2027 = (
        scenario.growth_start_2027 if growth_start_2027 is None else growth_start_2027
    )

    rows = []
    revenue = REV_2026_USDM
    previous_revenue = REV_2025_USDM
    nol = 0.0
    for i, year in enumerate(range(2026, 2036)):
        if year == 2026:
            growth = REV_2026_USDM / REV_2025_USDM - 1
        else:
            growth = growth_start_2027 + (scenario.growth_end_2035 - growth_start_2027) * (i - 1) / 8
            revenue = previous_revenue * (1 + growth)

        ebit_margin = -0.30 + (terminal_margin + 0.30) * i / 9
        ebit = revenue * ebit_margin
        beginning_nol = nol
        if ebit < 0:
            cash_tax = 0.0
            nol = beginning_nol - ebit
        else:
            nol_used = min(beginning_nol, ebit)
            taxable_income = ebit - nol_used
            cash_tax = taxable_income * TAX_RATE
            nol = beginning_nol - nol_used

        nopat = ebit - cash_tax
        reinvestment = max(revenue - previous_revenue, 0) / scenario.sales_to_capital
        fcff = nopat - reinvestment
        discount_factor = 1 / (1 + wacc) ** (i + 1)
        rows.append(
            {
                "year": year,
                "revenue": revenue,
                "growth_pct": growth * 100,
                "ebit_margin_pct": ebit_margin * 100,
                "ebit": ebit,
                "cash_tax": cash_tax,
                "nol_end": nol,
                "nopat": nopat,
                "reinvest": reinvestment,
                "fcff": fcff,
                "discount_factor": discount_factor,
                "pv_fcff": fcff * discount_factor,
            }
        )
        previous_revenue = revenue

    terminal_value = rows[-1]["fcff"] * (1 + TERMINAL_G) / (wacc - TERMINAL_G)
    pv_terminal = terminal_value * rows[-1]["discount_factor"]
    enterprise_value = sum(row["pv_fcff"] for row in rows) + pv_terminal
    equity_value = enterprise_value + NET_CASH_USDM
    per_share_hkd = equity_value / SHARES_M * USD_HKD

    return {
        "rows": rows,
        "terminal_value": terminal_value,
        "pv_terminal": pv_terminal,
        "enterprise_value": enterprise_value,
        "equity_value": equity_value,
        "per_share_hkd": per_share_hkd,
        "wacc": wacc,
        "terminal_margin": terminal_margin,
        "growth_start_2027": growth_start_2027,
    }


def write_base_projection_csv() -> None:
    base = project_scenario(SCENARIOS[1])["rows"]
    out = pd.DataFrame(base)[
        ["year", "revenue", "growth_pct", "ebit_margin_pct", "cash_tax", "nol_end", "nopat", "reinvest", "fcff"]
    ]
    for col in out.columns:
        if col != "year":
            out[col] = out[col].round(1)
    write_csv(out, EVENTSTUDY / "base_projection.csv")


def write_nonparametric_robustness_csv() -> None:
    """Event-level bootstrap and exact sign-test diagnostics for the CAR robustness table."""
    df = pd.read_csv(EVENTSTUDY / "car_robustness.csv")
    rng = np.random.default_rng(2513)
    windows = [
        ("Mean-adjusted reaction [0,+1]", "react_mean"),
        ("Mean-adjusted drift [+2,+10]", "drift_mean"),
        ("Peer-adjusted reaction [0,+1]", "react_peer"),
        ("Peer-adjusted drift [+2,+10]", "drift_peer"),
    ]
    rows = []
    for label, col in windows:
        values = df[col].to_numpy(dtype=float)
        n = len(values)
        positives = int((values > 0).sum())
        sign_p_one_sided = sum(comb(n, k) for k in range(positives, n + 1)) / (2 ** n)
        bootstrap_means = rng.choice(values, size=(20000, n), replace=True).mean(axis=1)
        lo, hi = np.percentile(bootstrap_means, [2.5, 97.5])
        rows.append(
            {
                "window": label,
                "n_events": n,
                "positive_events": positives,
                "mean_car_pct": round(float(values.mean()), 1),
                "median_car_pct": round(float(np.median(values)), 1),
                "sign_test_p_one_sided": round(float(sign_p_one_sided), 4),
                "bootstrap_mean_ci_2_5_pct": round(float(lo), 1),
                "bootstrap_mean_ci_97_5_pct": round(float(hi), 1),
            }
        )
    write_csv(pd.DataFrame(rows), EVENTSTUDY / "nonparametric_robustness.csv")


def write_block_bootstrap_outputs(n_boot: int = 20000) -> None:
    """Block-bootstrap peer-adjusted CAR means under a no-event null."""
    zhipu = pd.read_csv(ROOT / "data" / "Zhipu_KnowledgeAtlas_daily.csv")
    peer = pd.read_csv(ROOT / "data" / "MiniMax_daily.csv")
    zhipu["date"] = pd.to_datetime(zhipu["trade_date"].astype(str))
    peer["date"] = pd.to_datetime(peer["trade_date"].astype(str))
    merged = (
        zhipu[["date", "pct_chg"]]
        .rename(columns={"pct_chg": "zhipu_ret"})
        .merge(peer[["date", "pct_chg"]].rename(columns={"pct_chg": "peer_ret"}), on="date", how="inner")
        .sort_values("date")
        .reset_index(drop=True)
    )
    merged["peer_adjusted_ar"] = merged["zhipu_ret"] - merged["peer_ret"]

    events = [pd.Timestamp(day) for _, day in CAPABILITY_EVENTS]
    event_locs = [int(merged.index[merged["date"] == event_day][0]) for event_day in events]
    actual = pd.read_csv(EVENTSTUDY / "car_robustness.csv")
    actual_react_mean = float(actual["react_peer"].mean())
    actual_drift_mean = float(actual["drift_peer"].mean())
    drift_lengths = [9 if bool(full) else 7 for full in actual["drift_full"]]

    excluded_dates = set()
    for loc, drift_len in zip(event_locs, drift_lengths):
        for rel in range(0, 2):
            if 0 <= loc + rel < len(merged):
                excluded_dates.add(merged.loc[loc + rel, "date"])
        for rel in range(2, 2 + drift_len):
            if 0 <= loc + rel < len(merged):
                excluded_dates.add(merged.loc[loc + rel, "date"])

    rng = np.random.default_rng(2513)

    def valid_block_starts(length: int) -> np.ndarray:
        starts = []
        for start in range(0, len(merged) - length + 1):
            dates = set(merged.loc[start : start + length - 1, "date"])
            if not dates.intersection(excluded_dates):
                starts.append(start)
        return np.array(starts, dtype=int)

    def sampled_block_cars(lengths: list[int]) -> np.ndarray:
        draws = np.empty(n_boot)
        starts_by_length = {length: valid_block_starts(length) for length in sorted(set(lengths))}
        for length, starts in starts_by_length.items():
            if len(starts) == 0:
                raise ValueError(f"No valid block starts for length {length}")
        for i in range(n_boot):
            cars = []
            for length in lengths:
                start = int(rng.choice(starts_by_length[length]))
                cars.append(float(merged.loc[start : start + length - 1, "peer_adjusted_ar"].sum()))
            draws[i] = float(np.mean(cars))
        return draws

    react_draws = sampled_block_cars([2, 2, 2, 2])
    drift_draws = sampled_block_cars(drift_lengths)
    distribution = pd.DataFrame(
        {
            "draw": np.arange(1, n_boot + 1),
            "peer_reaction_mean_car_pct": react_draws,
            "peer_drift_mean_car_pct": drift_draws,
        }
    )
    write_csv(distribution, EVENTSTUDY / "block_bootstrap_distribution.csv")

    summary_rows = []
    for window, draws, actual_mean in [
        ("Peer-adjusted reaction [0,+1]", react_draws, actual_react_mean),
        ("Peer-adjusted drift [+2,+10]", drift_draws, actual_drift_mean),
    ]:
        lo, hi = np.percentile(draws, [2.5, 97.5])
        percentile = (np.sum(draws <= actual_mean) + 0.5 * np.sum(draws == actual_mean)) / len(draws) * 100
        summary_rows.append(
            {
                "window": window,
                "n_bootstrap": n_boot,
                "actual_mean_car_pct": round(actual_mean, 2),
                "bootstrap_mean_pct": round(float(np.mean(draws)), 2),
                "bootstrap_ci_2_5_pct": round(float(lo), 2),
                "bootstrap_ci_97_5_pct": round(float(hi), 2),
                "actual_percentile": round(float(percentile), 2),
                "block_lengths_days": ",".join(str(length) for length in ([2, 2, 2, 2] if "reaction" in window else drift_lengths)),
                "excluded_event_window_days": len(excluded_dates),
            }
        )
    write_csv(pd.DataFrame(summary_rows), EVENTSTUDY / "block_bootstrap_summary.csv")

    fig, axes = plt.subplots(1, 2, figsize=(10.8, 4.1), dpi=150)
    panels = [
        (axes[0], react_draws, actual_react_mean, summary_rows[0], "Reaction [0,+1]"),
        (axes[1], drift_draws, actual_drift_mean, summary_rows[1], "Drift [+2,+10]"),
    ]
    for ax, draws, actual_mean, summary, title in panels:
        lo = summary["bootstrap_ci_2_5_pct"]
        hi = summary["bootstrap_ci_97_5_pct"]
        percentile = summary["actual_percentile"]
        ax.hist(draws, bins=42, color=C_BLUE, alpha=0.78, edgecolor="white", linewidth=0.4)
        ax.axvspan(lo, hi, color=C_BLUE, alpha=0.14, label="95% null interval")
        ax.axvline(actual_mean, color=C_RED, linewidth=2.0, label="Actual mean CAR")
        ax.axvline(0, color=C_INK, linestyle=":", linewidth=1.0)
        ax.set_title(title)
        ax.set_xlabel("Bootstrapped mean CAR (%)")
        ax.set_ylabel("Frequency")
        ax.text(
            0.03,
            0.95,
            f"CI [{lo:.1f}, {hi:.1f}]%\nactual {actual_mean:.1f}%\npercentile {percentile:.1f}",
            transform=ax.transAxes,
            va="top",
            ha="left",
            fontsize=8.3,
            bbox={"boxstyle": "round,pad=0.25", "facecolor": "white", "edgecolor": C_GRAY, "alpha": 0.92},
        )
        ax.grid(axis="y", alpha=0.2)
        ax.spines[["top", "right"]].set_visible(False)
    axes[0].legend(loc="lower left", fontsize=7.8, framealpha=0.92)
    fig.suptitle("Block-bootstrap null distribution of peer-adjusted mean CAR", y=1.02)
    fig.tight_layout()
    savefig(fig, FIGURES / "fig10_block_bootstrap.png", bbox_inches="tight")
    plt.close(fig)


def write_workbook() -> None:
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Assumptions"
    assumptions = [
        ("GLOBAL ASSUMPTIONS", None, None),
        ("WACC", WACC, None),
        ("Terminal growth g", TERMINAL_G, None),
        ("Tax rate", TAX_RATE, None),
        ("Shares outstanding (m)", SHARES_M, None),
        ("USD/HKD", USD_HKD, None),
        (
            "Net cash (US$m)",
            NET_CASH_USDM,
            "net IPO proceeds plus pre-IPO cash; preferred shares convert at IPO",
        ),
        ("Market date", MARKET.date.strftime("%Y-%m-%d"), "derived from data/Zhipu_KnowledgeAtlas_daily.csv"),
        ("Market price (HK$)", MARKET.price_hkd, None),
        ("2025 revenue (US$m)", REV_2025_USDM, None),
        ("2026E revenue (US$m)", REV_2026_USDM, None),
    ]
    for row in assumptions:
        ws.append(row)
    ws["A1"].font = Font(bold=True)

    for scenario in SCENARIOS:
        ws = wb.create_sheet(scenario.name)
        ws.append([f"DCF - {scenario.name} scenario"])
        ws.append(["growth start (2027)", scenario.growth_start_2027])
        ws.append(["growth end (2035)", scenario.growth_end_2035])
        ws.append(["start op margin", -0.30])
        ws.append(["terminal op margin", scenario.terminal_margin])
        ws.append(["sales-to-capital", scenario.sales_to_capital])
        ws.append([])
        headers = [
            "Year",
            "i",
            "growth",
            "revenue",
            "op margin",
            "EBIT",
            "cash tax",
            "NOL end",
            "NOPAT",
            "dRev",
            "reinvest",
            "FCFF",
            "disc",
            "PV",
        ]
        ws.append(headers)
        for i, year in enumerate(range(2026, 2036), start=0):
            r = 9 + i
            prev_r = r - 1
            if year == 2026:
                growth_formula = "=Assumptions!$B$11/Assumptions!$B$10-1"
                revenue_formula = "=Assumptions!$B$11"
                drev_formula = "=D9-Assumptions!$B$10"
                prior_nol = "0"
            else:
                growth_formula = f"=$B$2+($B$3-$B$2)*(B{r}-1)/8"
                revenue_formula = f"=D{prev_r}*(1+C{r})"
                drev_formula = f"=D{r}-D{prev_r}"
                prior_nol = f"H{prev_r}"
            ws.append(
                [
                    year,
                    i,
                    growth_formula,
                    revenue_formula,
                    f"=$B$4+($B$5-$B$4)*B{r}/9",
                    f"=D{r}*E{r}",
                    f"=IF(F{r}<0,0,MAX(F{r}-{prior_nol},0)*Assumptions!$B$4)",
                    f"=IF(F{r}<0,{prior_nol}-F{r},MAX({prior_nol}-F{r},0))",
                    f"=F{r}-G{r}",
                    drev_formula,
                    f"=MAX(J{r},0)/$B$6",
                    f"=I{r}-K{r}",
                    f"=1/(1+Assumptions!$B$2)^(B{r}+1)",
                    f"=L{r}*M{r}",
                ]
            )

        ws.append([])
        ws.append(["Sum PV (explicit)", "=SUM(N9:N18)"])
        ws.append(["Terminal value", "=L18*(1+Assumptions!$B$3)/(Assumptions!$B$2-Assumptions!$B$3)"])
        ws.append(["PV terminal", "=B21*M18"])
        ws.append(["Enterprise value", "=B20+B22"])
        ws.append(["Equity value (US$m)", "=B23+Assumptions!$B$7"])
        ws.append(["Per share (US$)", "=B24/Assumptions!$B$5"])
        ws.append(["Per share (HK$)", "=B25*Assumptions!$B$6"])

        for cell in ws[8]:
            cell.font = Font(bold=True)
            cell.fill = PatternFill("solid", fgColor="D9EAF7")

    ws = wb.create_sheet("Summary")
    ws.append(["VALUATION SUMMARY"])
    ws.append(["Scenario", "Prob", "Equity (US$m)", "Per share (HK$)"])
    for i, scenario in enumerate(SCENARIOS, start=3):
        ws.append([scenario.name, scenario.probability, f"={scenario.name}!B24", f"={scenario.name}!B26"])
    ws.append(["PROB-WEIGHTED", None, "=SUMPRODUCT(B3:B5,C3:C5)/SUM(B3:B5)", "=SUMPRODUCT(B3:B5,D3:D5)/SUM(B3:B5)"])
    ws.append(["Market price (HK$)", "=Assumptions!B9"])
    ws.append(["DCF value as % of market", "=D6/B7"])
    ws.append(["Market equity value / revenue", "=(Assumptions!B9*Assumptions!B5/Assumptions!B6)/Assumptions!B11"])

    audit = wb.create_sheet("Audit Summary")
    results = {scenario.name: project_scenario(scenario) for scenario in SCENARIOS}
    weighted_equity = sum(s.probability * results[s.name]["equity_value"] for s in SCENARIOS)
    weighted_share = sum(s.probability * results[s.name]["per_share_hkd"] for s in SCENARIOS)
    audit_rows = [
        ("Market date", MARKET.date.strftime("%Y-%m-%d")),
        ("Market price (HK$)", round(MARKET.price_hkd, 1)),
        ("Market equity value (US$m)", round(MARKET.equity_value_usdm, 1)),
        ("Market EV / FY2026E revenue", round(MARKET.revenue_multiple, 1)),
        ("Bear per share (HK$)", round(results["Bear"]["per_share_hkd"], 1)),
        ("Base per share (HK$)", round(results["Base"]["per_share_hkd"], 1)),
        ("Bull per share (HK$)", round(results["Bull"]["per_share_hkd"], 1)),
        ("Probability-weighted equity (US$m)", round(weighted_equity, 1)),
        ("Probability-weighted per share (HK$)", round(weighted_share, 1)),
        ("DCF value as % of market", round(weighted_share / MARKET.price_hkd, 4)),
    ]
    for row in audit_rows:
        audit.append(row)

    summary_df = pd.DataFrame(
        [{"metric": metric, "value": value} for metric, value in audit_rows]
    )
    write_csv(summary_df, EVENTSTUDY / "valuation_summary.csv")

    for sheet in wb.worksheets:
        for col in range(1, sheet.max_column + 1):
            sheet.column_dimensions[get_column_letter(col)].width = 16
        sheet.freeze_panes = "A2"

    wb.calculation.fullCalcOnLoad = True
    wb.calculation.forceFullCalc = True
    wb.save(MODEL)


def per_share_for_base(**kwargs: float) -> float:
    return project_scenario(SCENARIOS[1], **kwargs)["per_share_hkd"]


def write_sensitivity_chart() -> None:
    base = per_share_for_base()
    cases = [
        ("WACC", per_share_for_base(wacc=0.15), per_share_for_base(wacc=0.12), "15% / 12%"),
        (
            "Terminal margin",
            per_share_for_base(terminal_margin=0.23),
            per_share_for_base(terminal_margin=0.33),
            "23% / 33%",
        ),
        (
            "2027 revenue growth",
            per_share_for_base(growth_start_2027=0.46),
            per_share_for_base(growth_start_2027=0.66),
            "46% / 66%",
        ),
    ]

    fig, ax = plt.subplots(figsize=(8.2, 3.6), dpi=150)
    y = np.arange(len(cases))
    for i, (label, low, high, range_label) in enumerate(cases):
        left = min(low, high)
        width = abs(high - low)
        ax.barh(i, width, left=left, height=0.45, color=C_BLUE)
        ax.plot([base, base], [i - 0.34, i + 0.34], color=C_INK, linewidth=1.2)
        ax.text(max(low, high) + 1.0, i, f"{range_label}: HK${low:.0f}-{high:.0f}", va="center", fontsize=8)
    ax.set_yticks(y)
    ax.set_yticklabels([case[0] for case in cases])
    ax.invert_yaxis()
    ax.set_xlabel("DCF value per share (HK$)")
    ax.set_title("Base-case DCF sensitivity")
    ax.axvline(base, color=C_INK, linewidth=1, linestyle="--")
    ax.text(base + 0.35, -0.48, f"Base HK${base:.0f}", fontsize=8, va="center")
    ax.grid(axis="x", alpha=0.25)
    ax.spines[["top", "right"]].set_visible(False)
    fig.tight_layout()
    savefig(fig, FIGURES / "fig7_sensitivity.png", bbox_inches="tight")
    plt.close(fig)


def write_football_field() -> None:
    results = {s.name: project_scenario(s) for s in SCENARIOS}
    weighted = sum(s.probability * results[s.name]["per_share_hkd"] for s in SCENARIOS)
    rows = [
        ("Scenario DCF", results["Bear"]["per_share_hkd"], results["Bull"]["per_share_hkd"]),
        ("Valuation / revenue", 30 * REV_2026_USDM / SHARES_M * USD_HKD, 40 * REV_2026_USDM / SHARES_M * USD_HKD),
        ("Market", MARKET.price_hkd, MARKET.price_hkd),
    ]
    fig, ax = plt.subplots(figsize=(8.6, 3.6), dpi=150)
    y = np.arange(len(rows))
    for i, (label, low, high) in enumerate(rows):
        ax.hlines(i, low, high, color=(C_RED if label == "Market" else C_BLUE), linewidth=9)
        ax.text(low, i + 0.18, f"HK${low:.0f}", ha="center", fontsize=8)
        if high != low:
            ax.text(high, i + 0.18, f"HK${high:.0f}", ha="center", fontsize=8)
    ax.scatter([weighted], [0], marker="D", color=C_ORANGE, zorder=3, label="Prob.-weighted DCF")
    comp_mid = 35 * REV_2026_USDM / SHARES_M * USD_HKD
    ax.scatter([comp_mid], [1], marker="D", color=C_GREEN, zorder=3, label="Multiple midpoint")
    ax.set_xscale("log")
    ax.set_yticks(y)
    ax.set_yticklabels([row[0] for row in rows])
    ax.invert_yaxis()
    ax.set_xlabel("HK$ per share (log scale)")
    ax.set_title("Valuation football field")
    ax.grid(axis="x", which="both", alpha=0.25)
    ax.spines[["top", "right", "left"]].set_visible(False)
    ax.legend(loc="lower right", fontsize=8)
    fig.tight_layout()
    savefig(fig, FIGURES / "fig5_football_field.png", bbox_inches="tight")
    plt.close(fig)


def write_comps_chart() -> None:
    names = ["MiniMax", "OpenAI / Anthropic", "Zhipu"]
    values = [35, 35, MARKET.revenue_multiple]
    colors = [C_BLUE, C_GREEN, C_RED]
    fig, ax = plt.subplots(figsize=(7.8, 4.0), dpi=150)
    ax.bar(names, values, color=colors)
    ax.set_yscale("log")
    ax.set_ylabel("Equity value / revenue (x, log scale)")
    ax.set_title("Revenue multiple comparison")
    for i, v in enumerate(values):
        ax.text(i, v * 1.12, f"{v:.0f}x", ha="center", fontsize=9)
    ax.grid(axis="y", which="both", alpha=0.25)
    ax.spines[["top", "right"]].set_visible(False)
    fig.tight_layout()
    savefig(fig, FIGURES / "fig6_ps_comps.png", bbox_inches="tight")
    plt.close(fig)


def write_financial_profile() -> None:
    """Company-introduction figure: revenue compounding vs cloud-margin compression."""
    labels = ["FY2022", "FY2023", "FY2024", "H1 2025"]
    x = np.arange(len(labels))
    revenue = np.array([57.4, 124.5, 312.4, 190.9])  # RMB millions
    total_margin = np.array([54.6, 64.6, 56.3, 50.0])  # %, FY2025 point is H1 2025
    cloud_margin = np.array([76.1, 31.0, 3.4, -0.4])  # %, FY2025 point is H1 2025

    fig, ax1 = plt.subplots(figsize=(8.2, 4.2), dpi=150)
    rev_mask = ~np.isnan(revenue)
    ax1.bar(x[rev_mask], revenue[rev_mask], width=0.55, color=C_BLUE, zorder=2,
            label="Revenue (RMB m)")
    for xi, rv in zip(x[rev_mask], revenue[rev_mask]):
        ax1.text(xi, rv + 18, f"{rv:.0f}", ha="center", fontsize=8, color=C_BLUE)
    ax1.set_ylabel("Revenue (RMB millions)", color=C_BLUE)
    ax1.tick_params(axis="y", labelcolor=C_BLUE)
    ax1.set_ylim(0, 840)
    ax1.set_xticks(x)
    ax1.set_xticklabels(labels)
    ax1.spines[["top"]].set_visible(False)

    ax2 = ax1.twinx()
    ax2.plot(x, total_margin, "o-", color=C_GREEN, linewidth=2,
             zorder=3, label="Total gross margin (%)")
    ax2.plot(x, cloud_margin, "o--", color=C_RED, linewidth=2,
             zorder=3, label="Cloud gross margin (%)")
    for xi, mg in zip(x, cloud_margin):
        ax2.text(xi, mg - 8, f"{mg:.1f}%", ha="center", fontsize=8, color=C_RED)
    ax2.axhline(0, color=C_GRAY, linewidth=0.7, linestyle=":")
    ax2.set_ylabel("Gross margin (%)", color=C_RED)
    ax2.tick_params(axis="y", labelcolor=C_RED)
    ax2.set_ylim(-12, 100)
    ax2.spines[["top"]].set_visible(False)
    ax2.legend(loc="upper right", fontsize=7.5, framealpha=0.9)

    ax1.set_title("Revenue scales while cloud deployment margins compress")
    fig.tight_layout()
    savefig(fig, FIGURES / "fig8_financial_profile.png", bbox_inches="tight")
    plt.close(fig)


def write_glm_timeline() -> None:
    """Company-introduction figure: milestones from founding (2019) to GLM-5.2 (2026).

    Events are evenly spaced (not on a proportional date axis) so the dense 2026 release
    cadence stays legible alongside the longer 2019--2025 build-up.
    """
    events = [
        ("2019", "Founded — Tsinghua KEG", C_TEAL),
        ("2022", "GLM family · cloud GM 76%", C_TEAL),
        ("2024", "Revenue RMB 312m", C_TEAL),
        ("H1 2025", "Revenue RMB 191m · total GM 50%", C_TEAL),
        ("8 Jan 2026", "IPO HK$116.20 · Ch.18C", C_GREEN),
        ("11 Feb 2026", "GLM-5", C_ORANGE),
        ("16 Mar 2026", "GLM-5-Turbo", C_ORANGE),
        ("8 Apr 2026", "GLM-5.1 · #1 SWE-Bench Pro", C_ORANGE),
        ("5–8 Jun 2026", "HSTECH index · Stock Connect", C_GREEN),
        ("15 Jun 2026", "GLM-5.2 · MIT weights · 1M ctx", C_RED),
    ]
    n = len(events)
    fig, ax = plt.subplots(figsize=(13.0, 3.6), dpi=150)
    ax.axhline(0, color="#444444", linewidth=1.4, zorder=1)
    for i, (date_lbl, name, color) in enumerate(events):
        y = 0.62 if i % 2 == 0 else -0.62
        ax.plot([i, i], [0, y], color=color, linewidth=1.1, zorder=1)
        ax.scatter([i], [0], s=70, color=color, zorder=3)
        ax.annotate(
            f"{date_lbl}\n{name}", (i, y), ha="center",
            va="bottom" if y > 0 else "top", fontsize=8.2, color="#222222",
        )
    ax.set_xlim(-0.6, n - 0.4)
    ax.set_ylim(-1.55, 1.55)
    ax.set_xticks([])
    ax.set_yticks([])
    ax.set_title("From a Tsinghua lab (2019) to GLM-5.2 (2026): the path to and beyond the IPO")
    ax.spines[["top", "right", "left", "bottom"]].set_visible(False)
    fig.tight_layout()
    savefig(fig, FIGURES / "fig9_glm_timeline.png", bbox_inches="tight")
    plt.close(fig)


# Event taxonomy shared by the price/return figures: capability releases vs index/flow catalysts.
CAPABILITY_EVENTS = [
    ("GLM-5", "2026-02-11"),
    ("GLM-5-Turbo", "2026-03-16"),
    ("GLM-5.1", "2026-04-08"),
    ("GLM-5.2", "2026-06-15"),
]
FLOW_EVENTS = [
    ("IPO", "2026-01-08"),
    ("HSTECH exp.", "2026-02-20"),
    ("Connect exp.", "2026-05-13"),
    ("HSTECH in", "2026-06-05"),
    ("Connect in", "2026-06-08"),
]
CAP_COLOR = C_ORANGE   # capability-release markers
FLOW_COLOR = C_GREEN   # index/flow markers
ZHIPU_COLOR = C_RED    # protagonist series
MINIMAX_COLOR = C_BLUE  # peer series


def _load_prices(filename: str, ipo_price: float) -> pd.DataFrame:
    df = pd.read_csv(ROOT / "data" / filename)
    df["date"] = pd.to_datetime(df["trade_date"].astype(str))
    df = df.sort_values("date").reset_index(drop=True)
    df["rebased"] = df["close"] / ipo_price * 100.0
    return df


def write_price_paths() -> None:
    """Rebased price paths (IPO=100, log) with capability and index/flow event markers."""
    from matplotlib.lines import Line2D

    zhipu = _load_prices("Zhipu_KnowledgeAtlas_daily.csv", 116.2)
    mmx = _load_prices("MiniMax_daily.csv", 165.0)
    fig, ax = plt.subplots(figsize=(9.8, 4.9), dpi=150)
    ax.plot(zhipu["date"], zhipu["rebased"], color=ZHIPU_COLOR, linewidth=1.7,
            label="Zhipu / Knowledge Atlas (2513.HK)")
    ax.plot(mmx["date"], mmx["rebased"], color=MINIMAX_COLOR, linewidth=1.7,
            label="MiniMax (00100.HK)")
    ax.set_yscale("log")
    lo = min(zhipu["rebased"].min(), mmx["rebased"].min())
    hi = max(zhipu["rebased"].max(), mmx["rebased"].max())
    ax.set_ylim(lo * 0.7, hi * 1.9)
    for name, day in CAPABILITY_EVENTS:
        x = pd.Timestamp(day)
        ax.axvline(x, color=CAP_COLOR, linestyle="--", linewidth=1.0, alpha=0.9, zorder=1)
        ax.text(x, hi * 1.95, name, rotation=90, va="top", ha="center", fontsize=6.6,
                color=CAP_COLOR)
    for name, day in FLOW_EVENTS:
        x = pd.Timestamp(day)
        ax.axvline(x, color=FLOW_COLOR, linestyle=":", linewidth=1.0, alpha=0.85, zorder=1)
        ax.text(x, lo * 0.74, name, rotation=90, va="bottom", ha="center", fontsize=6.6,
                color=FLOW_COLOR)
    ax.set_ylabel("Index (IPO = 100, log scale)")
    ax.set_title("Rebased price paths: capability releases and index/flow events")
    ax.grid(axis="y", which="both", alpha=0.2)
    ax.spines[["top", "right"]].set_visible(False)
    handles = [
        Line2D([], [], color=ZHIPU_COLOR, linewidth=1.7, label="Zhipu / Knowledge Atlas (2513.HK)"),
        Line2D([], [], color=MINIMAX_COLOR, linewidth=1.7, label="MiniMax (00100.HK)"),
        Line2D([], [], color=CAP_COLOR, linestyle="--", label="Capability release"),
        Line2D([], [], color=FLOW_COLOR, linestyle=":", label="Index / flow event"),
    ]
    ax.legend(handles=handles, loc="lower right", fontsize=7.5, framealpha=0.9)
    fig.tight_layout()
    savefig(fig, FIGURES / "fig1_price_paths.png", bbox_inches="tight")
    plt.close(fig)


def write_daily_returns() -> None:
    """Zhipu daily returns with capability- and flow-event days highlighted and labelled."""
    zhipu = _load_prices("Zhipu_KnowledgeAtlas_daily.csv", 116.2)
    cap = {pd.Timestamp(d): n for n, d in CAPABILITY_EVENTS}
    flow = {pd.Timestamp(d): n for n, d in FLOW_EVENTS}
    fig, ax = plt.subplots(figsize=(10.6, 4.7), dpi=150)
    colors = []
    for d in zhipu["date"]:
        if d in cap:
            colors.append(C_ORANGE)
        elif d in flow:
            colors.append(C_GREEN)
        else:
            colors.append(C_GRAY)
    ax.bar(zhipu["date"], zhipu["pct_chg"], color=colors, width=1.6, zorder=2)
    ax.axhline(0, color=C_INK, linewidth=0.8)
    for d, r in zip(zhipu["date"], zhipu["pct_chg"]):
        name = cap.get(d) or flow.get(d)
        if name:
            ax.text(d, r + (1.6 if r >= 0 else -1.6), name, rotation=90,
                    ha="center", va="bottom" if r >= 0 else "top", fontsize=6.6,
                    color=C_ORANGE if d in cap else C_GREEN)
    ax.set_ylabel("Daily return (%)")
    ax.set_title("Zhipu daily returns: capability releases vs index/flow events")
    ax.set_ylim(zhipu["pct_chg"].min() - 10, zhipu["pct_chg"].max() + 12)
    ax.grid(axis="y", alpha=0.2)
    ax.spines[["top", "right"]].set_visible(False)
    from matplotlib.patches import Patch
    ax.legend(handles=[
        Patch(color=C_ORANGE, label="Capability release"),
        Patch(color=C_GREEN, label="Index / flow event"),
        Patch(color=C_GRAY, label="Other day"),
    ], loc="lower left", fontsize=7.5, framealpha=0.9)
    fig.tight_layout()
    savefig(fig, FIGURES / "fig2_daily_returns.png", bbox_inches="tight")
    plt.close(fig)


def event_car_series() -> dict[str, pd.Series]:
    prices = pd.read_csv(ROOT / "data" / "Zhipu_KnowledgeAtlas_daily.csv")
    prices["trade_date"] = pd.to_datetime(prices["trade_date"].astype(str))
    prices["ret"] = prices["close"].pct_change()
    events = {
        "GLM-5": "2026-02-11",
        "GLM-5-Turbo": "2026-03-16",
        "GLM-5.1": "2026-04-08",
        "GLM-5.2": "2026-06-15",
    }
    series = {}
    for name, date in events.items():
        idx = prices.index[prices["trade_date"] == pd.Timestamp(date)][0]
        expected = prices.loc[idx - 20 : idx - 6, "ret"].mean()
        rel_days = []
        ars = []
        for rel in range(-5, 11):
            loc = idx + rel
            if 0 <= loc < len(prices):
                rel_days.append(rel)
                ars.append((prices.loc[loc, "ret"] - expected) * 100)
        series[name] = pd.Series(ars, index=rel_days).cumsum()
    return series


def write_event_chart() -> None:
    series = event_car_series()
    full_days = [day for day in range(-5, 11) if all(day in s.index for s in series.values())]
    avg = pd.concat([s.loc[full_days] for s in series.values()], axis=1).mean(axis=1)
    fig, ax = plt.subplots(figsize=(8.5, 5.0), dpi=150)
    event_colors = [C_BLUE, C_ORANGE, C_GREEN, C_RED]
    for (name, s), shade in zip(series.items(), event_colors):
        ax.plot(s.index, s.values, color=shade, alpha=0.9, linewidth=1.3, label=name)
    ax.plot(avg.index, avg.values, color=C_INK, linewidth=2.6, label="Average (n=4)")
    ax.axvline(0, color=C_GRAY, linestyle="--", linewidth=1)
    ax.axhline(0, color=C_GRAY, linewidth=0.8)
    ax.set_xlim(-5, 10)
    ax.set_xlabel("Event day")
    ax.set_ylabel("CAR (%)")
    ax.set_title("Average CAR in event time")
    ax.text(8.1, ax.get_ylim()[0] + 2, "n falls after +8", fontsize=8)
    ax.legend(fontsize=8, ncol=2, loc="upper left")
    ax.grid(alpha=0.25)
    ax.spines[["top", "right"]].set_visible(False)
    fig.tight_layout()
    savefig(fig, FIGURES / "fig3_car_eventtime.png", bbox_inches="tight")
    plt.close(fig)


def write_reaction_vs_drift() -> None:
    """fig4: immediate reaction vs post-event drift (mean-adjusted), matching Table 2."""
    glm = [
        ("GLM-5", 22.5, 24.7),
        ("GLM-5-Turbo", 7.7, -19.3),
        ("GLM-5.1", 13.8, -14.2),
        ("GLM-5.2 †", 30.8, 28.2),
    ]
    minimax = ("MiniMax M2.7", -5.5, -49.1)
    fig, ax = plt.subplots(figsize=(7.4, 5.0), dpi=150)
    ax.axhline(0, color=C_GRAY, linewidth=1)
    ax.axvline(0, color=C_GRAY, linewidth=1)
    for name, rx, dy in glm:
        ax.scatter([rx], [dy], s=62, color=C_RED, zorder=3)
        ax.annotate(name, (rx, dy), textcoords="offset points", xytext=(8, 4),
                    fontsize=8.5, color=C_RED)
    ax.scatter([minimax[1]], [minimax[2]], s=62, facecolor=C_BLUE, edgecolor=C_INK,
               linewidth=1.0, zorder=3)
    ax.annotate(minimax[0], (minimax[1], minimax[2]), textcoords="offset points",
                xytext=(8, 4), fontsize=8.5, color=C_BLUE)
    ax.text(0.02, 0.97, "UNDER-reaction\n(drift continues)", transform=ax.transAxes,
            va="top", ha="left", fontsize=8, color=C_GREEN)
    ax.text(0.02, 0.52, "OVER-reaction\n(reversal)", transform=ax.transAxes,
            va="top", ha="left", fontsize=8, color=C_RED)
    ax.set_xlabel("Immediate reaction CAR[0,+1] (%)")
    ax.set_ylabel("Post-event drift CAR[+2,+10] (%)")
    ax.set_title("Reaction vs. Drift (mean-adjusted; GLM-5.2 drift provisional †)")
    ax.grid(alpha=0.2)
    ax.spines[["top", "right"]].set_visible(False)
    fig.tight_layout()
    savefig(fig, FIGURES / "fig4_reaction_vs_drift.png", bbox_inches="tight")
    plt.close(fig)


def main() -> None:
    write_price_summary_csv()
    write_base_projection_csv()
    write_event_panel_outputs(write_csv)
    write_nonparametric_robustness_csv()
    write_block_bootstrap_outputs()
    write_workbook()
    write_sensitivity_chart()
    write_football_field()
    write_comps_chart()
    write_financial_profile()
    write_glm_timeline()
    write_price_paths()
    write_daily_returns()
    write_event_chart()
    write_reaction_vs_drift()


if __name__ == "__main__":
    main()

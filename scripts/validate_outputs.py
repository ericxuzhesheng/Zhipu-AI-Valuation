from __future__ import annotations

import argparse
import csv
import re
import subprocess
import sys
from pathlib import Path

import openpyxl
import pandas as pd
from pypdf import PdfReader


ROOT = Path(__file__).resolve().parents[1]
PAPER = ROOT / "paper" / "main.tex"
SUBMISSION_GLOB = "42353012_*.pdf"


def fail(message: str) -> None:
    raise AssertionError(message)


def submission_pdf_path() -> Path:
    matches = sorted(ROOT.glob(SUBMISSION_GLOB))
    if len(matches) != 1:
        found = ", ".join(path.name for path in matches) or "none"
        fail(f"expected exactly one root-level {SUBMISSION_GLOB} file; found {found}")
    return matches[0]


def texcount_words() -> int:
    proc = subprocess.run(
        ["texcount", "-inc", "-sum", "main.tex"],
        cwd=ROOT / "paper",
        text=True,
        capture_output=True,
        check=False,
    )
    if proc.returncode != 0:
        fail(f"texcount failed: {proc.stderr.strip() or proc.stdout.strip()}")
    match = re.search(r"Words in text:\s+(\d+)", proc.stdout)
    if not match:
        fail("texcount output did not include 'Words in text'")
    return int(match.group(1))


def abstract_words() -> int:
    text = PAPER.read_text(encoding="utf-8")
    match = re.search(r"\\begin\{abstract\}(.*?)\\end\{abstract\}", text, re.S)
    if not match:
        fail("abstract environment not found")
    abstract = match.group(1)
    abstract = re.sub(r"\\[a-zA-Z]+\*?(?:\[[^\]]*\])?(?:\{([^{}]*)\})?", r" \1 ", abstract)
    abstract = re.sub(r"\$[^$]*\$", " ", abstract)
    abstract = re.sub(r"[^A-Za-z0-9'-]+", " ", abstract)
    return len([word for word in abstract.split() if re.search(r"[A-Za-z0-9]", word)])


def check_price_summary_dates() -> None:
    summary = pd.read_csv(ROOT / "data" / "price_summary.csv")
    specs = {
        "02513.HK": ROOT / "data" / "Zhipu_KnowledgeAtlas_daily.csv",
        "00100.HK": ROOT / "data" / "MiniMax_daily.csv",
        "01956.HK": ROOT / "data" / "WengeAI_daily.csv",
    }
    for code, path in specs.items():
        raw = pd.read_csv(path)
        expected = int(raw["trade_date"].max())
        actual_rows = summary.loc[summary["code"] == code, "latest_date"]
        if actual_rows.empty:
            fail(f"price_summary.csv missing {code}")
        actual = int(actual_rows.iloc[0])
        if actual != expected:
            fail(f"{code} latest_date mismatch: summary {actual}, raw {expected}")


def check_valuation_summary() -> None:
    csv_path = ROOT / "eventstudy" / "valuation_summary.csv"
    if not csv_path.exists():
        fail("eventstudy/valuation_summary.csv missing")
    with csv_path.open(newline="", encoding="utf-8") as f:
        csv_rows = {row["metric"]: row["value"] for row in csv.DictReader(f)}
    wb = openpyxl.load_workbook(ROOT / "model" / "valuation_model.xlsx", data_only=False, read_only=True)
    if "Audit Summary" not in wb.sheetnames:
        fail("model workbook missing Audit Summary sheet")
    sheet_rows = {
        str(metric): str(value)
        for metric, value in wb["Audit Summary"].iter_rows(min_row=1, max_col=2, values_only=True)
        if metric is not None
    }
    for metric, value in csv_rows.items():
        if metric not in sheet_rows:
            fail(f"Audit Summary missing metric {metric}")
        sheet_value = sheet_rows[metric]
        if value != sheet_value:
            try:
                if abs(float(value) - float(sheet_value)) <= 1e-9:
                    continue
            except ValueError:
                pass
            fail(f"valuation_summary mismatch for {metric}: csv {value}, workbook {sheet_value}")


def check_pdfs() -> None:
    for path in [ROOT / "paper" / "main.pdf", submission_pdf_path()]:
        if not path.exists():
            fail(f"PDF missing: {path}")
        if path.stat().st_size < 100_000:
            fail(f"PDF too small: {path}")
        pages = len(PdfReader(path).pages)
        if pages <= 0:
            fail(f"PDF has no pages: {path}")


def check_event_panel() -> None:
    panel = ROOT / "eventstudy" / "event_panel.csv"
    catalog = ROOT / "eventstudy" / "event_catalog.csv"
    summary = ROOT / "eventstudy" / "event_panel_summary.csv"
    input_catalog = ROOT / "data" / "event_catalog_input.csv"
    for path in [input_catalog, panel, catalog, summary]:
        if not path.exists():
            fail(f"event-study output missing: {path}")
    panel_rows = pd.read_csv(panel)
    if len(panel_rows) < 9:
        fail(f"event panel has too few computable events: {len(panel_rows)}")
    catalog_rows = pd.read_csv(catalog)
    input_rows = pd.read_csv(input_catalog, keep_default_na=False)
    if len(catalog_rows) < len(panel_rows):
        fail("event catalog should include all panel rows and excluded candidates")
    if len(catalog_rows) != len(input_rows):
        fail(f"event catalog row count {len(catalog_rows)} does not match input {len(input_rows)}")


def check_appendix_outputs() -> None:
    required = [
        ROOT / "data" / "comps_beta_bridge_input.csv",
        ROOT / "data" / "comps_beta_bridge.csv",
        ROOT / "paper" / "beta_bridge_auto.tex",
        ROOT / "eventstudy" / "reverse_dcf_sensitivity.csv",
        ROOT / "figures" / "fig11_reverse_dcf_heatmap.png",
    ]
    for path in required:
        if not path.exists():
            fail(f"appendix output missing: {path}")
        if path.stat().st_size <= 0:
            fail(f"appendix output is empty: {path}")
    beta = pd.read_csv(ROOT / "data" / "comps_beta_bridge.csv")
    if len(beta) < 5:
        fail(f"comps beta bridge has too few rows: {len(beta)}")
    reverse = pd.read_csv(ROOT / "eventstudy" / "reverse_dcf_sensitivity.csv")
    if len(reverse) < 20:
        fail(f"reverse DCF grid has too few rows: {len(reverse)}")


def main() -> int:
    parser = argparse.ArgumentParser(description="Validate generated project outputs.")
    parser.add_argument(
        "--skip-tex-checks",
        action="store_true",
        help="Skip checks that require TeX tooling or freshly compiled PDFs.",
    )
    args = parser.parse_args()

    checks = [
        ("price_summary dates match raw data", check_price_summary_dates),
        ("valuation_summary matches Excel Audit Summary", check_valuation_summary),
        ("event panel outputs exist", check_event_panel),
        ("appendix outputs exist", check_appendix_outputs),
    ]
    if not args.skip_tex_checks:
        checks = [
            ("texcount text words <= 3000", lambda: texcount_words() <= 3000),
            ("abstract words < 400", lambda: abstract_words() < 400),
            ("PDF files exist and have pages", check_pdfs),
            *checks,
        ]
    for label, check in checks:
        result = check()
        if result is False:
            fail(label)
        print(f"OK: {label}")
    return 0


if __name__ == "__main__":
    try:
        raise SystemExit(main())
    except AssertionError as exc:
        print(f"FAIL: {exc}", file=sys.stderr)
        raise SystemExit(1)
